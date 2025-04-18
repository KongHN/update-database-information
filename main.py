import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from tqdm import tqdm


def get_column_index(df, column_name):
    """
    获取指定列名在 DataFrame 中的索引
    :param df: DataFrame
    :param column_name: 列名
    :return: 列索引，如果未找到则返回 None
    """
    col_indices = df.columns[df.iloc[0] == column_name]
    return col_indices[0] if not col_indices.empty else None


def set_cell_style(cell, sample_font, sample_alignment):
    """
    设置单元格的字体和对齐方式
    :param cell: 单元格对象
    :param sample_font: 示例字体
    :param sample_alignment: 示例对齐方式
    :return: 无
    """
    cell.font = Font(name=sample_font.name, size=sample_font.size,
                     bold=sample_font.bold, italic=sample_font.italic,
                     color=sample_font.color)
    cell.alignment = Alignment(horizontal=sample_alignment.horizontal,
                               vertical=sample_alignment.vertical)


def update_cell(ws, index, col_index, value, sample_font, sample_alignment):
    """
    更新单元格的值，并设置字体和对齐方式
    :param ws: 工作表对象
    :param index: 行索引
    :param col_index: 列索引
    :param value: 单元格的值
    :param sample_font: 示例字体
    :param sample_alignment: 示例对齐方式
    :return: 无
    """
    if col_index is not None:
        cell = ws.cell(row=index, column=col_index + 1)
        cell.value = value
        set_cell_style(cell, sample_font, sample_alignment)


def process_csv_files(csv_folder_path, xlsx_file_path):
    """
    该函数用于处理 CSV 文件和 XLSX 文件的匹配及数据更新操作。
    :param csv_folder_path: CSV 文件所在文件夹路径
    :param xlsx_file_path: XLSX 文件路径
    :return: 无
    """
    try:
        wb = load_workbook(xlsx_file_path)
        ws = wb.active
        xlsx_df = pd.DataFrame(ws.values)
    except FileNotFoundError:
        print(f"错误: 未找到 {xlsx_file_path} 文件。")
        return

    # 获取第一列第二行单元格的字体格式和对齐方式
    sample_cell = ws.cell(row=2, column=1)
    sample_font = sample_cell.font
    sample_alignment = sample_cell.alignment

    csv_files = [file for file in os.listdir(csv_folder_path) if file.endswith('.csv')]
    total_files = len(csv_files)

    with tqdm(total=total_files, desc="处理 CSV 文件进度") as pbar:
        for file in csv_files:
            # 从 CSV 文件名中提取数据名称
            data_name = file.split('data_')[-1].replace('.csv', '')
            # 获取 Name_in_database 列的索引
            name_in_database_col_index = get_column_index(xlsx_df, 'Name_in_database')
            if name_in_database_col_index is None:
                continue
            # 在 XLSX 文件中查找匹配的数据行索引
            match_index = xlsx_df[xlsx_df.iloc[:, name_in_database_col_index] == data_name].index
            if not match_index.empty:
                # 将 DataFrame 索引转换为 Excel 行号（从 1 开始）
                index = match_index[0] + 1
                try:
                    csv_df = pd.read_csv(os.path.join(csv_folder_path, file))
                except FileNotFoundError:
                    continue

                def check_accuracy(alt_columns):
                    """
                    检查 CSV 文件中是否存在 Accuracy 列，并返回相应的判断结果。
                    :param alt_columns: 备用列名列表（在此处未使用）
                    :return: 'yes' 如果 CSV 文件存在 Accuracy 列，否则 'no'
                    """
                    return 'yes' if any(col in csv_df.columns for col in alt_columns) else 'no'

                def check_num_blocks(alt_columns):
                    """
                    检查 CSV 文件中是否存在与块相关的列，并返回块的唯一值数量或 'no'。
                    :param alt_columns: 与块相关的列名列表
                    :return: 块的唯一值数量，如果不存在相关列则返回 'no'
                    """
                    block_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    return 'no' if block_col is None else csv_df[block_col].nunique()

                def check_rt_confidence(alt_columns):
                    """
                    检查 CSV 文件中是否存在与反应时间置信度相关的列，并返回相应的判断结果。
                    :param alt_columns: 与反应时间置信度相关的列名列表
                    :return: 'yes' 如果 CSV 文件存在相关列，否则 'no'
                    """
                    confidence_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    return 'yes' if confidence_col is not None else 'no'

                def check_blank_values(alt_columns):
                    """
                    检查 CSV 文件中是否存在空白值、NaN 或 NAN，并返回相应的判断结果。
                    :param alt_columns: 备用列名列表（在此处未使用）
                    :return: 'yes' 如果 CSV 文件存在空白值、NaN 或 NAN，否则 'no'
                    """
                    return 'yes' if csv_df.isnull().values.any() else 'no'

                def check_block_min_max(alt_columns):
                    """
                    检查 CSV 文件中是否存在与块相关的列，并统计每个 Subj_idx 中该列不同值的个数，返回最多和最少个数，若不存在相关列则返回 ('no', 'no')。
                    :param alt_columns: 与块相关的列名列表
                    :return: 包含最多和最少数的元组，如果不存在相关列则返回 ('no', 'no')
                    """
                    block_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    if block_col is None or 'Subj_idx' not in csv_df.columns:
                        return 'no', 'no'
                    unique_counts = csv_df.groupby('Subj_idx')[block_col].nunique()
                    if unique_counts.empty:
                        return 'no', 'no'
                    max_count = unique_counts.max()
                    min_count = unique_counts.min()
                    return max_count, min_count

                def check_stimulus():
                    """
                    检查 CSV 文件中是否存在列名包含 Stimulus 的列
                    :return: 'yes' 如果存在，否则 'no'
                    """
                    return 'yes' if any('Stimulus' in col for col in csv_df.columns) else 'no'

                def check_response():
                    """
                    检查 CSV 文件中是否存在列名包含 Response 的列
                    :return: 'yes' 如果存在，否则 'no'
                    """
                    return 'yes' if any('Response' in col for col in csv_df.columns) else 'no'

                def check_trial_in_block(alt_columns):
                    """
                    先按 Subj_idx 列分组，再统计每个分组里 block 列相同值的数量，返回最大和最小数量。
                    :param alt_columns: 与块相关的列名列表
                    :return: 包含最大和最小数量的元组，如果不存在相关列则返回 ('no', 'no')
                    """
                    block_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    if block_col is None or 'Subj_idx' not in csv_df.columns:
                        return 'no', 'no'
                    all_counts = []
                    for subj, group in csv_df.groupby('Subj_idx'):
                        block_counts = group[block_col].value_counts()
                        all_counts.extend(block_counts.tolist())
                    if not all_counts:
                        return 'no', 'no'
                    max_count = max(all_counts)
                    min_count = min(all_counts)
                    return max_count, min_count

                def check_same_num_blocks(alt_columns):
                    block_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    if block_col is None or 'Subj_idx' not in csv_df.columns:
                        return 'no'
                    unique_counts = csv_df.groupby('Subj_idx')[block_col].nunique()
                    return 'yes' if len(set(unique_counts)) == 1 else 'no'

                def check_same_num_trial_in_block(alt_columns):
                    block_col = next((col for col in alt_columns if col in csv_df.columns), None)
                    if block_col is None:
                        return 'no'
                    try:
                        grouped_by_subj = csv_df.groupby('Subj_idx')
                        for _, subj_group in grouped_by_subj:
                            subj_block_grouped = subj_group.groupby(block_col)
                            counts = [len(group) for _, group in subj_block_grouped]
                            if len(set(counts)) != 1:
                                return 'no'
                        return 'yes'
                    except KeyError:
                        return 'no'

                accuracy_columns = ['Accuracy', 'Accuracy_col', 'Accuracy_let', 'ErrorDirection',
                                    'ErrorDirectionJudgment', 'Accuracy_Motion', 'Accuracy_Color']
                block_columns = ['block', 'blocks', 'Block', 'Blocks', 'BlockNumber', 'Block_count',
                                 'Int.Block', 'block_type', 'BlockID', 'Block_Type', 'NumBlock', 'blocki']
                confidence_columns = ['RT_confidence', 'RT_conf', 'RT_decConf', 'RT_decConf_1', 'RT_decConf_2']
                trial_in_block_columns = ['Trial_in_block', 'trials_per_block', 'NumTrialinBlock', 'Trial_count',
                                          'Trial in Block', 'Trial_number', 'trial_count', 'Trial_Number', 'Trial']

                columns_to_check = [
                    ('Accuracy', accuracy_columns, check_accuracy),
                    ('Num_Blocks', block_columns, check_num_blocks),
                    ('RT_Confidence', confidence_columns, check_rt_confidence),
                    ('Blank_Value', [], check_blank_values),
                    ('Stimulus', [], check_stimulus),
                    ('Response', [], check_response)
                ]

                for col_name, alt_columns, func in columns_to_check:
                    col_index = get_column_index(xlsx_df, col_name)
                    if func.__code__.co_argcount == 1:
                        value = func(alt_columns)
                    else:
                        value = func()
                    update_cell(ws, index, col_index, value, sample_font, sample_alignment)

                max_trial, min_trial = check_trial_in_block(block_columns)
                update_cell(ws, index, get_column_index(xlsx_df, 'Max_Trial_in_Block'), max_trial, sample_font,
                            sample_alignment)
                update_cell(ws, index, get_column_index(xlsx_df, 'Min_Trial_in_Block'), min_trial, sample_font,
                            sample_alignment)

                max_block, min_block = check_block_min_max(block_columns)
                update_cell(ws, index, get_column_index(xlsx_df, 'Max_Num_Blocks'), max_block, sample_font,
                            sample_alignment)
                update_cell(ws, index, get_column_index(xlsx_df, 'Min_Num_Blocks'), min_block, sample_font,
                            sample_alignment)

                same_num_blocks = check_same_num_blocks(block_columns)
                update_cell(ws, index, get_column_index(xlsx_df, 'Same_Num_Blocks'), same_num_blocks, sample_font,
                            sample_alignment)

                same_num_trial_in_block = check_same_num_trial_in_block(block_columns)
                # 注意这里要使用实际的列名，包含空格
                update_cell(ws, index, get_column_index(xlsx_df, 'Same_Num_Trial_in_Block'), same_num_trial_in_block,
                            sample_font, sample_alignment)

            pbar.update(1)

    try:
        wb.save(xlsx_file_path)
        print("文件更新成功。")
    except Exception as e:
        print(f"错误: 保存文件时出现问题: {e}")


if __name__ == "__main__":
    csv_folder_path = input("请输入 CSV 文件所在文件夹的路径: ")
    xlsx_file_path = input("请输入要更新的 XLSX 文件的路径: ")
    process_csv_files(csv_folder_path, xlsx_file_path)